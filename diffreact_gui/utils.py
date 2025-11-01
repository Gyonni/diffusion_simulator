from __future__ import annotations

import json
import logging
import os
from dataclasses import asdict
from typing import Any, Dict

import numpy as np

from .config import RESULTS_DIR, LOG_LEVEL
from .models import SimParams


def setup_logging(level: str = LOG_LEVEL) -> None:
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )


def validate_params(p: SimParams) -> None:
    if not p.layers:
        raise ValueError("At least one layer must be specified")
    total_length = 0.0
    for idx, layer in enumerate(p.layers):
        if layer.thickness <= 0:
            raise ValueError(f"Layer {idx + 1} thickness must be positive")
        if layer.diffusivity <= 0:
            raise ValueError(f"Layer {idx + 1} diffusivity must be positive")
        if layer.reaction_rate < 0:
            raise ValueError(f"Layer {idx + 1} reaction rate must be non-negative")
        if layer.nodes < 2:
            raise ValueError(f"Layer {idx + 1} nodes must be >= 2")
        total_length += layer.thickness
    if total_length <= 0:
        raise ValueError("Total stack thickness must be positive")
    if p.dt <= 0:
        raise ValueError("dt must be positive")
    if p.t_max <= 0:
        raise ValueError("t_max must be positive")
    if p.Cs < 0:
        raise ValueError("Cs must be non-negative")
    if p.bc_right not in {"Neumann", "Dirichlet"}:
        raise ValueError("bc_right must be one of: Neumann, Dirichlet")


def ensure_results_dir(path: str = RESULTS_DIR) -> str:
    os.makedirs(path, exist_ok=True)
    return path


def cumulative_trapz(y: np.ndarray, x: np.ndarray) -> np.ndarray:
    """Return cumulative trapezoidal integral of y(x).

    Parameters
    ----------
    y : np.ndarray
        Shape (N,) array of function values
    x : np.ndarray
        Shape (N,) array, must be strictly increasing

    Returns
    -------
    np.ndarray
        Cumulative integral array of same shape as y

    Raises
    ------
    ValueError
        If y and x shapes don't match or if x is not strictly increasing
    """
    if y.shape != x.shape:
        raise ValueError(f"Shape mismatch: y.shape={y.shape}, x.shape={x.shape}")
    if len(x) > 1 and np.any(np.diff(x) <= 0):
        raise ValueError("x array must be strictly increasing")

    out = np.zeros_like(y)
    if len(y) <= 1:
        return out

    # Vectorized computation for better performance
    dx = np.diff(x)
    avg_y = 0.5 * (y[1:] + y[:-1])
    out[1:] = np.cumsum(avg_y * dx)
    return out


def save_results_npz(
    base_path: str,
    *,
    t: np.ndarray,
    x: np.ndarray,
    C_xt: np.ndarray,
    J_source: np.ndarray,
    J_target: np.ndarray,
    J_end: np.ndarray,
    cum_source: np.ndarray,
    cum_target: np.ndarray,
    cum_end: np.ndarray,
    mass_target: np.ndarray,
    layer_boundaries: np.ndarray,
    D_nodes: np.ndarray | None = None,
    D_edges: np.ndarray | None = None,
    J_probe: np.ndarray | None = None,
    cum_probe: np.ndarray | None = None,
    probe_position: float | None = None,
) -> str:
    npz_path = os.path.join(base_path, "results.npz")
    payload: Dict[str, Any] = {
        "t": t,
        "x": x,
        "C_xt": C_xt,
        "J_source": J_source,
        "J_target": J_target,
        "J_end": J_end,
        "cum_source": cum_source,
        "cum_target": cum_target,
        "cum_end": cum_end,
        "mass_target": mass_target,
        "layer_boundaries": layer_boundaries,
    }
    if D_nodes is not None:
        payload["D_nodes"] = D_nodes
    if D_edges is not None:
        payload["D_edges"] = D_edges
    if J_probe is not None:
        payload["J_probe"] = J_probe
    if cum_probe is not None:
        payload["cum_probe"] = cum_probe
    if probe_position is not None:
        payload["probe_position"] = probe_position
    np.savez(npz_path, **payload)
    return npz_path


def save_csv_flux(
    base_path: str,
    t: np.ndarray,
    J_source: np.ndarray,
    J_target: np.ndarray,
    J_end: np.ndarray,
    cum_source: np.ndarray,
    cum_target: np.ndarray,
    cum_end: np.ndarray,
    mass_target: np.ndarray,
    *,
    J_probe: np.ndarray | None = None,
    cum_probe: np.ndarray | None = None,
    filename: str = "flux_vs_time.csv",
) -> str:
    csv_path = os.path.join(base_path, filename)
    columns = [
        t,
        J_source,
        J_target,
        J_end,
        cum_source,
        cum_target,
        cum_end,
        mass_target,
    ]
    header = [
        "t[s]",
        "Flux_surface[mol/(m^2*s)]",
        "Flux_target_interface[mol/(m^2*s)]",
        "Flux_exit[mol/(m^2*s)]",
        "Cum_flux_surface[mol/m^2]",
        "Cum_flux_target_interface[mol/m^2]",
        "Cum_flux_exit[mol/m^2]",
        "Mass_target[mol/m^2]",
    ]
    if J_probe is not None and J_probe.size:
        columns.append(J_probe)
        header.append("Flux_probe[mol/(m^2*s)]")
    if cum_probe is not None and cum_probe.size:
        columns.append(cum_probe)
        header.append("Cum_flux_probe[mol/m^2]")
    data = np.column_stack(columns)
    header_line = ",".join(header)
    np.savetxt(csv_path, data, delimiter=",", header=header_line, comments="")
    return csv_path


def save_csv_profile(base_path: str, x: np.ndarray, C: np.ndarray, t_value: float) -> str:
    csv_path = os.path.join(base_path, f"profile_t_{t_value:.6g}s.csv")
    header = "x[m],C[mol/m^3]"
    data = np.column_stack([x, C])
    np.savetxt(csv_path, data, delimiter=",", header=header, comments="")
    return csv_path


def save_profiles_matrix(base_path: str, x: np.ndarray, t: np.ndarray, C_xt: np.ndarray, filename: str = "concentration_profiles.csv") -> str:
    csv_path = os.path.join(base_path, filename)
    header = ["x[m]"] + [f"C(t={ti:.6g}s)[mol/m^3]" for ti in t]
    data = np.column_stack([x, C_xt.T])
    np.savetxt(csv_path, data, delimiter=",", header=",".join(header), comments="")
    return csv_path


def save_metadata(base_path: str, params: SimParams, extras: Dict[str, Any] | None = None) -> str:
    meta_path = os.path.join(base_path, "metadata.json")
    payload: Dict[str, Any] = asdict(params)
    if extras:
        payload.update(extras)
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    return meta_path


# Material Library Functions
MATERIALS_LIBRARY_FILE = "materials_library.json"


def load_materials_library(library_path: str = MATERIALS_LIBRARY_FILE) -> Dict[str, Dict[str, float]]:
    """Load materials library from JSON file.

    Returns:
        Dictionary mapping material name to properties dict with keys:
        'D0' (optional), 'Ea' (optional), 'diffusivity' (optional), 'reaction_rate'
    """
    if not os.path.exists(library_path):
        return {}
    try:
        with open(library_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        logging.warning(f"Failed to load materials library: {e}")
        return {}


def save_materials_library(materials: Dict[str, Dict[str, float]], library_path: str = MATERIALS_LIBRARY_FILE) -> None:
    """Save materials library to JSON file.

    Args:
        materials: Dictionary mapping material name to properties dict
        library_path: Path to save the library file
    """
    try:
        with open(library_path, "w", encoding="utf-8") as f:
            json.dump(materials, f, indent=2)
    except IOError as e:
        logging.error(f"Failed to save materials library: {e}")
        raise


def add_material_to_library(
    name: str,
    D0: float | None = None,
    Ea: float | None = None,
    diffusivity: float | None = None,
    reaction_rate: float = 0.0,
    library_path: str = MATERIALS_LIBRARY_FILE
) -> None:
    """Add or update a material in the library.

    Args:
        name: Material name
        D0: Pre-exponential factor [m^2/s]
        Ea: Activation energy [eV]
        diffusivity: Direct diffusivity [m^2/s]
        reaction_rate: Reaction rate constant [1/s]
        library_path: Path to the library file
    """
    materials = load_materials_library(library_path)
    material_data: Dict[str, float] = {"reaction_rate": reaction_rate}
    if D0 is not None:
        material_data["D0"] = D0
    if Ea is not None:
        material_data["Ea"] = Ea
    if diffusivity is not None:
        material_data["diffusivity"] = diffusivity
    materials[name] = material_data
    save_materials_library(materials, library_path)


def save_temperature_sweep_excel(
    base_path: str,
    temperatures: np.ndarray,
    results_by_temp: Dict[float, Dict[str, Any]],
    filename: str = "results_temperature_sweep.xlsx"
) -> str:
    """Save temperature sweep results as Excel file with temperature-labeled sheets.

    Args:
        base_path: Directory where the file will be saved
        temperatures: Array of temperatures [K]
        results_by_temp: Dictionary mapping temperature to simulation results
        filename: Name of the output Excel file

    Returns:
        str: Full path to the saved file
    """
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for i_temp, T in enumerate(temperatures):
        temp_result = results_by_temp[T]
        sheet_name = f"{int(T)}K"

        # Flux data sheet
        ws_flux = wb.create_sheet(title=f"{sheet_name}_Flux")
        t = temp_result["t"]
        J_source = temp_result["J_source"]
        J_target = temp_result["J_target"]
        J_end = temp_result["J_end"]
        cum_source = temp_result["cum_source"]
        cum_target = temp_result["cum_target"]
        cum_end = temp_result["cum_end"]
        mass_target = temp_result["mass_target"]

        # Headers
        headers = [
            "Time [s]",
            "Flux_surface [mol/(m^2·s)]",
            "Flux_target_interface [mol/(m^2·s)]",
            "Flux_exit [mol/(m^2·s)]",
            "Cum_flux_surface [mol/m^2]",
            "Cum_flux_target_interface [mol/m^2]",
            "Cum_flux_exit [mol/m^2]",
            "Mass_target [mol/m^2]",
        ]

        # Add probe columns if available
        J_probe_temp = temp_result.get("J_probe")
        cum_probe_temp = temp_result.get("cum_probe")
        if J_probe_temp is not None and J_probe_temp.size > 0:
            headers.extend(["Flux_probe [mol/(m^2·s)]", "Cum_flux_probe [mol/m^2]"])

        ws_flux.append(headers)

        # Data rows
        for i in range(len(t)):
            row = [
                t[i],
                J_source[i],
                J_target[i],
                J_end[i],
                cum_source[i],
                cum_target[i],
                cum_end[i],
                mass_target[i],
            ]
            if J_probe_temp is not None and J_probe_temp.size > 0:
                row.extend([J_probe_temp[i], cum_probe_temp[i]])
            ws_flux.append(row)

        # Concentration profile sheet
        ws_conc = wb.create_sheet(title=f"{sheet_name}_Concentration")
        x = temp_result["x"]
        C_xt = temp_result["C_xt"]

        # Excel has a maximum of 16,384 columns (XFD)
        # If we have too many time steps, transpose the data structure
        MAX_EXCEL_COLS = 16384
        n_time = len(t)
        n_x = len(x)

        if n_time + 1 > MAX_EXCEL_COLS:
            # Transpose: rows are times, columns are positions
            # Headers: Time, then x positions
            conc_headers = ["Time [s]"] + [f"C(x={x[i]:.6e}m) [mol/m^3]" for i in range(n_x)]
            ws_conc.append(conc_headers)

            # Data rows: each row is one time with concentrations at all positions
            for i_t in range(n_time):
                row = [t[i_t]] + [C_xt[i_t, i_x] for i_x in range(n_x)]
                ws_conc.append(row)
        else:
            # Original structure: rows are positions, columns are times
            # Headers: Position, then time steps
            conc_headers = ["Position x [m]"] + [f"C(t={t[i]:.6e}s) [mol/m^3]" for i in range(n_time)]
            ws_conc.append(conc_headers)

            # Data rows: each row is one position with concentrations at all times
            for i_x in range(n_x):
                row = [x[i_x]] + [C_xt[i_t, i_x] for i_t in range(n_time)]
                ws_conc.append(row)

    # Save workbook
    full_path = os.path.join(base_path, filename)
    wb.save(full_path)
    return full_path
